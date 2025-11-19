from __future__ import annotations

from dataclasses import dataclass
import logging
import os
from pathlib import Path
from typing import Any, Dict, List, Optional
import json

import openpyxl


logger = logging.getLogger(__name__)

import pandas as pd


# Custom exceptions
class InvalidExcelError(Exception):
    """Raised when an uploaded or existing Excel file is invalid or cannot be opened."""


ALLOWED_EXCEL_EXTS = {".xlsx", ".xlsm", ".xltx", ".xltm"}


def load_workbook_safe(path: str, **kwargs) -> openpyxl.Workbook:
    """Load an Excel workbook with safety checks.

    Validates that the file exists, has a supported extension and isn't trivially small
    (< 512 bytes). Wraps any load errors in :class:`InvalidExcelError`.
    """
    p = Path(path)
    if not p.exists():
        raise InvalidExcelError("Excel file not found")

    if p.suffix.lower() not in ALLOWED_EXCEL_EXTS:
        raise InvalidExcelError("Unsupported Excel extension")

    try:
        size = p.stat().st_size
    except OSError as e:
        raise InvalidExcelError("Cannot read Excel file size") from e

    if size <= 512:
        raise InvalidExcelError("Excel file is too small or empty")

    try:
        wb = openpyxl.load_workbook(p, **kwargs)
        return wb
    except Exception as e:
        raise InvalidExcelError(f"Failed to open Excel workbook: {e}") from e


def save_workbook_atomic(workbook: openpyxl.Workbook, path: str) -> None:
    """Save workbook atomically: write to a tmp file then replace the target path.

    This prevents partial writes from leaving corrupted files in place.
    """
    tmp_path = f"{path}.tmp"
    workbook.save(tmp_path)
    os.replace(tmp_path, path)


@dataclass
class Condition:
    column: str
    value: Any


def parse_simple_condition(condition: Optional[str]) -> Optional[Condition]:
    """
    Parse a very simple SQL-like condition of the form 'column=value'.
    This is intentionally constrained and should be extended carefully.
    """
    if not condition:
        return None
    if "=" not in condition:
        return None
    column, value = condition.split("=", 1)
    column = column.strip()
    value = value.strip().strip("'").strip('"')
    return Condition(column=column, value=value)


def _ensure_excel_path(path: Path) -> Path:
    """
    Normalize and validate the Excel path.
    """
    resolved = path.resolve()
    if not os.path.exists(resolved):
        raise FileNotFoundError("Excel file not found")
    allowed = {".xlsx", ".xlsm", ".xltx", ".xltm"}
    if resolved.suffix.lower() not in allowed:
        raise ValueError("Only .xlsx/.xlsm/.xltx/.xltm Excel files are supported.")

    logger.info(
        "Opening Excel file",
        extra={
            "path": str(resolved),
            "extension": resolved.suffix.lower(),
            "size_bytes": os.path.getsize(resolved),
        },
    )
    return resolved


def _get_sheet(
    wb: openpyxl.Workbook, sheet_name: Optional[str]
) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Return the requested sheet or the first sheet if sheet_name is None.
    Raise a clean error if the sheet does not exist.
    """
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found in Excel file.")
        return wb[sheet_name]
    return wb[wb.sheetnames[0]]


def excel_read(
    path: Path, sheet_name: Optional[str], condition: Optional[str]
) -> List[Dict[str, Any]]:
    """
    Read Excel data using openpyxl, with optional simple condition.
    """
    resolved = _ensure_excel_path(path)
    wb = load_workbook_safe(str(resolved), data_only=True)
    try:
        sheet = _get_sheet(wb, sheet_name)

        header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            return []

        headers = [cell.value for cell in header_row]
        header_index = {name: idx for idx, name in enumerate(headers)}
        cond = parse_simple_condition(condition)

        rows: List[Dict[str, Any]] = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = {headers[i]: row[i] for i in range(len(headers))}
            if cond and cond.column in header_index:
                if str(record.get(cond.column)) != str(cond.value):
                    continue
            rows.append(record)
        return rows
    finally:
        wb.close()


def _normalise_insert_values(
    sheet: openpyxl.worksheet.worksheet.Worksheet, values: Any
) -> List[Any]:
    """
    Normalise INSERT values into a list aligned with the sheet header.

    Supported input formats:
    - list/tuple: used directly (padded/truncated to header length)
    - dict: mapped by header names
    - JSON string: parsed and processed recursively
    - comma-separated string: split and processed as list
    """
    header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
    if not header_row:
        # No headers; only accept a flat list as-is
        if isinstance(values, (list, tuple)):
            return list(values)
        raise ValueError(
            "Sheet has no header row; INSERT VALUES must be a list in column order."
        )

    headers = [cell.value for cell in header_row]

    # If dict: map by header names
    if isinstance(values, dict):
        return [values.get(h) for h in headers]

    # If already list/tuple: align length
    if isinstance(values, (list, tuple)):
        row = list(values)
        if len(row) < len(headers):
            row.extend([None] * (len(headers) - len(row)))
        elif len(row) > len(headers):
            row = row[: len(headers)]
        return row

    # If string: try JSON first, then comma-separated list
    if isinstance(values, str):
        text = values.strip()
        if text:
            try:
                parsed = json.loads(text)
                return _normalise_insert_values(sheet, parsed)
            except json.JSONDecodeError:
                parts = [p.strip() for p in text.split(",")]
                return _normalise_insert_values(sheet, parts)

    raise ValueError(
        "INSERT VALUES must be a list, dict, JSON string, or comma-separated string."
    )


def excel_insert(path: Path, sheet_name: Optional[str], values: Any) -> None:
    resolved = _ensure_excel_path(path)
    wb = load_workbook_safe(str(resolved))
    try:
        sheet = _get_sheet(wb, sheet_name)
        normalised = _normalise_insert_values(sheet, values)
        sheet.append(normalised)
        wb.save(resolved)
    finally:
        wb.close()


def excel_update(path: Path, sheet_name: str, condition: str, values: Dict[str, Any]) -> int:
    resolved = _ensure_excel_path(path)
    wb = load_workbook_safe(str(resolved))
    try:
        sheet = _get_sheet(wb, sheet_name)
        cond = parse_simple_condition(condition)
        if not cond:
            return 0

        header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            return 0

        headers = [cell.value for cell in header_row]
        header_index = {name: idx for idx, name in enumerate(headers)}
        if cond.column not in header_index:
            return 0

        match_count = 0
        for row in sheet.iter_rows(min_row=2):
            cell = row[header_index[cond.column]]
            if str(cell.value) == str(cond.value):
                for col_name, new_val in values.items():
                    if col_name in header_index:
                        row[header_index[col_name]].value = new_val
                match_count += 1

        wb.save(resolved)
        return match_count
    finally:
        wb.close()


def excel_delete(path: Path, sheet_name: str, condition: str) -> int:
    resolved = _ensure_excel_path(path)
    wb = load_workbook_safe(str(resolved))
    try:
        sheet = _get_sheet(wb, sheet_name)
        cond = parse_simple_condition(condition)
        if not cond:
            return 0

        header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            return 0

        headers = [cell.value for cell in header_row]
        header_index = {name: idx for idx, name in enumerate(headers)}
        if cond.column not in header_index:
            return 0

        rows_to_delete = []
        for idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            cell = row[header_index[cond.column]]
            if str(cell.value) == str(cond.value):
                rows_to_delete.append(idx)

        # Delete from bottom to top so indices remain valid
        for row_idx in reversed(rows_to_delete):
            sheet.delete_rows(row_idx)

        wb.save(resolved)
        return len(rows_to_delete)
    finally:
        wb.close()


def lookup_email_by_name(path: str, sheet: str, employee_name: str) -> Optional[str]:
    """Lookup an employee email by name from an Excel sheet using pandas.

    - Normalizes column names to lowercase and stripped.
    - Supports multiple header variants for name and email.
    - Returns the first matching email string or None if not found.
    """
    try:
        df = pd.read_excel(path, sheet_name=sheet, dtype=str)
    except Exception as e:
        logger.debug("Failed to read Excel file %s sheet %s: %s", path, sheet, e)
        return None

    # Normalize headers
    df.columns = df.columns.str.lower().str.strip()

    name_candidates = ["name", "employee_name", "full_name"]
    email_candidates = ["email", "e-mail", "email_address", "mail"]

    name_col = next((c for c in name_candidates if c in df.columns), None)
    email_col = next((c for c in email_candidates if c in df.columns), None)

    if not name_col or not email_col:
        return None

    target = employee_name.strip().lower()
    matches = df[df[name_col].astype(str).str.strip().str.lower() == target]

    if matches.empty:
        return None

    emails = matches[email_col].astype(str).tolist()
    if len(emails) > 1:
        logger.warning("Multiple email matches for name '%s' in %s:%s; returning first", employee_name, path, sheet)

    return emails[0]
